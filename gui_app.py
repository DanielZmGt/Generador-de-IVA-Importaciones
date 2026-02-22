import sys
import os
import glob
import pandas as pd
import numpy as np
import re
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl.styles import Font, PatternFill

# --- Logic: Convert ASC to XLSX ---
def detect_delimiter(filepath):
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            line = f.readline()
            if '|' in line: return '|'
            if '\t' in line: return '\t'
    except Exception:
        pass
    return '|'

def convert_asc_files(directory, log_callback=None):
    if log_callback: log_callback("Searching for .asc files...")
    
    pattern = os.path.join(directory, "*.asc")
    asc_files = glob.glob(pattern)
    
    if not asc_files:
        if log_callback: log_callback("No .asc files found.")
        return False, "No .asc files found."

    converted_count = 0
    for filepath in asc_files:
        filename = os.path.basename(filepath)
        output_filename = os.path.splitext(filename)[0] + ".xlsx"
        output_path = os.path.join(directory, output_filename)
        
        if log_callback: log_callback(f"Converting {filename}...")

        try:
            sep = detect_delimiter(filepath)
            df = pd.read_csv(filepath, sep=sep, dtype=str, encoding='utf-8', encoding_errors='ignore', skipinitialspace=True, index_col=False)
            
            df.columns = df.columns.str.strip()
            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].str.strip()

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Data')
                ws = writer.sheets['Data']
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
            
            converted_count += 1
        except Exception as e:
            if log_callback: log_callback(f"Error converting {filename}: {e}")
            return False, f"Error converting {filename}: {e}"
            
    if log_callback: log_callback(f"Successfully converted {converted_count} files.")
    return True, "Conversion complete."

# --- Logic: Generate IVA Report ---
def find_file(directory, suffix):
    # Search for *suffix.xlsx in the given directory
    pattern = os.path.join(directory, f"*{suffix}.xlsx")
    files = glob.glob(pattern)
    # Filter out result files to avoid circular dependency if re-running
    files = [f for f in files if "IVA_Importaciones" not in os.path.basename(f) and "ejemplo" not in os.path.basename(f)]
    
    if not files:
        return None
    return files[0]

def load_excel_as_str(filepath):
    try:
        df = pd.read_excel(filepath, sheet_name='Data', dtype=str)
        df.columns = df.columns.str.strip()
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].str.strip()
        return df
    except Exception as e:
        return None

def generate_iva_report(directory, log_callback=None):
    if log_callback: log_callback("Starting Report Generation...")
    
    # 1. Load 557
    file_557 = find_file(directory, "557")
    if not file_557:
        return False, "Critical: File with suffix 557 not found."
    
    if log_callback: log_callback(f"Loading base file: {os.path.basename(file_557)}")
    df_557 = load_excel_as_str(file_557)
    if df_557 is None:
        return False, "Failed to read 557 file."

    required_557 = ['Pedimento', 'Fraccion', 'SecuenciaFraccion', 'ClaveContribucion', 'FormaPago']
    if not all(col in df_557.columns for col in required_557):
        return False, f"557 file missing columns: {required_557}"

    # Filter
    mask = (df_557['ClaveContribucion'] == '3') & (df_557['FormaPago'] == '0')
    df_557 = df_557[mask].copy()
    if log_callback: log_callback(f"Filtered 557: {len(df_557)} rows match criteria.")

    df_557['Pedimento_Key'] = df_557['Pedimento'] + df_557['Fraccion'] + df_557['SecuenciaFraccion']

    # 2. Load 551
    file_551 = find_file(directory, "551")
    if file_551:
        if log_callback: log_callback(f"Merging with 551: {os.path.basename(file_551)}")
        df_551 = load_excel_as_str(file_551)
        if df_551 is not None and all(c in df_551.columns for c in ['Pedimento', 'Fraccion', 'SecuenciaFraccion']):
            df_551['Pedimento_Key'] = df_551['Pedimento'] + df_551['Fraccion'] + df_551['SecuenciaFraccion']
            
            map_551 = {
                'ClaveDocumento': 'Clave Pedimento',
                'ValorAduana': 'Valor Aduana',
                'ValorComercial': 'Valor Comercial',
                'DescripcionMercancia': 'Descripcion'
            }
            cols_to_pull = [c for c in map_551.keys() if c in df_551.columns]
            subset_551 = df_551[['Pedimento_Key'] + cols_to_pull].drop_duplicates('Pedimento_Key')
            df_557 = df_557.merge(subset_551, on='Pedimento_Key', how='left')
            df_557.rename(columns=map_551, inplace=True)

    # 3. Load 505
    file_505 = find_file(directory, "505")
    if file_505:
        if log_callback: log_callback(f"Merging with 505: {os.path.basename(file_505)}")
        df_505 = load_excel_as_str(file_505)
        if df_505 is not None and 'Pedimento' in df_505.columns:
            map_505 = {
                'ProveedorMercancia': 'Proveedor',
                'IndentFiscalProveedor': 'TaxID',
                'PaisFacturacion': 'País'
            }
            target_fields = list(map_505.keys())
            existing_fields = [c for c in target_fields if c in df_505.columns]
            
            if 'IndentFiscalProveedor' in df_505.columns:
                df_505['IndentFiscalProveedor'] = df_505['IndentFiscalProveedor'].apply(
                    lambda x: re.sub(r'[^a-zA-Z0-9]', '', str(x)) if pd.notna(x) else x
                )

            df_505[existing_fields] = df_505[existing_fields].replace('', np.nan)
            df_505['data_score'] = df_505[existing_fields].notna().sum(axis=1)
            df_505_sorted = df_505.sort_values(by=['Pedimento', 'data_score'], ascending=[True, False])
            subset_505 = df_505_sorted[['Pedimento'] + existing_fields].drop_duplicates('Pedimento', keep='first')
            
            df_557['Pedimento_Clean'] = df_557['Pedimento']
            df_557 = df_557.merge(subset_505, left_on='Pedimento_Clean', right_on='Pedimento', how='left')
            df_557.rename(columns=map_505, inplace=True)
            
            if 'Pedimento_y' in df_557.columns: df_557.drop(columns=['Pedimento_y'], inplace=True)
            if 'Pedimento_x' in df_557.columns: df_557.rename(columns={'Pedimento_x': 'Pedimento'}, inplace=True)
            if 'Pedimento_Clean' in df_557.columns: df_557.drop(columns=['Pedimento_Clean'], inplace=True)

    # 4. Load 701
    file_701 = find_file(directory, "701")
    rectified_in_col_b = set()
    
    if file_701:
        if log_callback: log_callback(f"Checking rectifications in 701: {os.path.basename(file_701)}")
        df_701 = load_excel_as_str(file_701)
        if df_701 is not None and 'Pedimento' in df_701.columns:
            rectified_in_col_b.update(df_701['Pedimento'].dropna().unique())

    df_557['rectificado'] = df_557['Pedimento'].apply(lambda x: x if x in rectified_in_col_b else 'NO')

    # Fill Country
    if 'Proveedor' in df_557.columns and 'País' in df_557.columns:
        valid_countries = df_557.dropna(subset=['Proveedor', 'País'])
        valid_countries = valid_countries[valid_countries['País'].str.strip() != '']
        if not valid_countries.empty:
            provider_country_map = valid_countries.set_index('Proveedor')['País'].to_dict()
            def fill_country(row):
                curr = str(row['País']).strip()
                if not curr or curr.lower() == 'nan':
                    return provider_country_map.get(row['Proveedor'], row['País'])
                return row['País']
            df_557['País'] = df_557.apply(fill_country, axis=1)

    # Final Columns
    target_cols = [
        'Patente', 'Pedimento', 'SeccionAduanera', 'Fraccion', 'SecuenciaFraccion', 
        'ClaveContribucion', 'FormaPago', 'ImportePago', 'FechaPagoReal', 
        'Clave Pedimento', 'Valor Aduana', 'Valor Comercial', 'Descripcion', 
        'Proveedor', 'TaxID', 'País', 'rectificado'
    ]
    for col in target_cols:
        if col not in df_557.columns:
            df_557[col] = ""
            
    df_final = df_557[target_cols].copy()
    
    if 'FechaPagoReal' in df_final.columns:
        df_final['FechaPagoReal'] = pd.to_datetime(df_final['FechaPagoReal'], dayfirst=True, errors='coerce')

    numeric_cols = ['Patente', 'Pedimento', 'SeccionAduanera', 'SecuenciaFraccion', 'ImportePago', 'Valor Aduana', 'Valor Comercial']
    for col in numeric_cols:
        if col in df_final.columns:
            df_final[col] = pd.to_numeric(df_final[col], errors='coerce')

    # Export
    output_file = os.path.join(directory, "IVA_Importaciones_Generado.xlsx")
    if log_callback: log_callback(f"Saving to {os.path.basename(output_file)}...")
    
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='IVA importaciones')
            ws = writer.sheets['IVA importaciones']
            
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            header_font = Font(bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            rect_idx = df_final.columns.get_loc('rectificado') + 1
            for row in ws.iter_rows(min_row=2):
                if str(row[rect_idx-1].value) != 'NO':
                    for cell in row:
                        cell.fill = red_fill
                        
        if log_callback: log_callback("Report generated successfully!")
        return True, "Success"
        
    except Exception as e:
        return False, f"Error saving file: {e}"

# --- GUI Application ---
class IvaGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Generador IVA Importaciones")
        self.root.geometry("600x450")
        
        # Set Icon
        if getattr(sys, 'frozen', False):
            # Running as compiled exe
            application_path = sys._MEIPASS
        else:
            application_path = os.path.dirname(os.path.abspath(__file__))
            
        icon_path = os.path.join(application_path, 'logo.ico')
        if os.path.exists(icon_path):
            self.root.iconbitmap(icon_path)

        # Style
        style = ttk.Style()
        style.theme_use('clam')
        
        # Header
        header_frame = ttk.Frame(root, padding="20")
        header_frame.pack(fill=tk.X)
        
        lbl_title = ttk.Label(header_frame, text="Generador de Reporte IVA Importaciones", font=("Helvetica", 16, "bold"))
        lbl_title.pack()
        
        lbl_subtitle = ttk.Label(header_frame, text="Convierte archivos .asc y genera el reporte Excel automáticamente.", font=("Helvetica", 10))
        lbl_subtitle.pack(pady=5)

        # Main Content
        content_frame = ttk.Frame(root, padding="20")
        content_frame.pack(fill=tk.BOTH, expand=True)

        # Directory Selection
        lbl_dir = ttk.Label(content_frame, text="1. Selecciona la carpeta con los archivos .asc:")
        lbl_dir.pack(anchor=tk.W)
        
        dir_frame = ttk.Frame(content_frame)
        dir_frame.pack(fill=tk.X, pady=5)
        
        self.entry_dir = ttk.Entry(dir_frame)
        self.entry_dir.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        btn_browse = ttk.Button(dir_frame, text="Examinar...", command=self.browse_directory)
        btn_browse.pack(side=tk.RIGHT)

        # Action Button
        self.btn_run = ttk.Button(content_frame, text="Generar Reporte", command=self.start_process, state=tk.DISABLED)
        self.btn_run.pack(pady=20, fill=tk.X)

        # Log Area
        lbl_log = ttk.Label(content_frame, text="Progreso:")
        lbl_log.pack(anchor=tk.W)
        
        self.txt_log = tk.Text(content_frame, height=10, state=tk.DISABLED, bg="#f0f0f0", font=("Consolas", 9))
        self.txt_log.pack(fill=tk.BOTH, expand=True)
        
        # Status Bar
        self.status_var = tk.StringVar()
        self.status_var.set("Listo")
        status_bar = ttk.Label(root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def browse_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.entry_dir.delete(0, tk.END)
            self.entry_dir.insert(0, directory)
            self.btn_run['state'] = tk.NORMAL
            self.log(f"Carpeta seleccionada: {directory}")

    def log(self, message):
        self.root.after(0, self._log_impl, message)

    def _log_impl(self, message):
        self.txt_log.config(state=tk.NORMAL)
        self.txt_log.insert(tk.END, message + "\n")
        self.txt_log.see(tk.END)
        self.txt_log.config(state=tk.DISABLED)

    def start_process(self):
        directory = self.entry_dir.get()
        if not os.path.isdir(directory):
            messagebox.showerror("Error", "La carpeta seleccionada no es válida.")
            return
            
        self.btn_run['state'] = tk.DISABLED
        self.log("-" * 40)
        
        # Run in a separate thread to keep GUI responsive
        thread = threading.Thread(target=self.run_logic, args=(directory,))
        thread.start()

    def run_logic(self, directory):
        try:
            self.status_var.set("Convirtiendo archivos...")
            success, msg = convert_asc_files(directory, self.log)
            if not success:
                self.log(f"Error: {msg}")
                messagebox.showerror("Error", msg)
                return

            self.status_var.set("Generando reporte...")
            success, msg = generate_iva_report(directory, self.log)
            
            if success:
                self.status_var.set("Completado")
                messagebox.showinfo("Éxito", "El reporte se ha generado correctamente en la misma carpeta.")
            else:
                self.status_var.set("Error")
                self.log(f"Error: {msg}")
                messagebox.showerror("Error", msg)

        except Exception as e:
            self.log(f"Excepción inesperada: {e}")
            messagebox.showerror("Error Crítico", str(e))
        finally:
            # Re-enable button on main thread
            self.root.after(0, lambda: self.btn_run.config(state=tk.NORMAL))

if __name__ == "__main__":
    root = tk.Tk()
    app = IvaGeneratorApp(root)
    root.mainloop()
