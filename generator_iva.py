import pandas as pd
import glob
import os
import sys
import numpy as np
import re
from openpyxl.styles import Font, PatternFill

def find_file(suffix):
    pattern = f"*{suffix}.xlsx"
    files = glob.glob(pattern)
    files = [f for f in files if "IVA_Importaciones" not in f and "ejemplo" not in f]
    if not files:
        print(f"Warning: File with suffix {suffix} not found.")
        return None
    return files[0]

def load_excel_as_str(filepath):
    print(f"Loading {filepath}...")
    try:
        df = pd.read_excel(filepath, sheet_name='Data', dtype=str)
        df.columns = df.columns.str.strip()
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].str.strip()
        return df
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
        return None

def main():
    print("Starting IVA Generator (from converted Excel files)...")
    
    # --- STEP 1: Load 557 (Base) ---
    file_557 = find_file("557")
    if not file_557:
        print("Critical: Base file 557 not found. Make sure convert_asc_to_xlsx.py has been run.")
        return

    df_557 = load_excel_as_str(file_557)
    if df_557 is None: return

    # Ensure critical columns exist
    required_557 = ['Pedimento', 'Fraccion', 'SecuenciaFraccion', 'ClaveContribucion', 'FormaPago']
    if not all(col in df_557.columns for col in required_557):
        print(f"Error: 557 file missing required columns: {required_557}")
        return

    # Filter: ClaveContribucion=3, FormaPago=0
    mask = (df_557['ClaveContribucion'] == '3') & (df_557['FormaPago'] == '0')
    df_557 = df_557[mask].copy()
    print(f"Filtered 557: {len(df_557)} rows remaining.")

    # Key
    df_557['Pedimento_Key'] = (
        df_557['Pedimento'] + 
        df_557['Fraccion'] + 
        df_557['SecuenciaFraccion']
    )

    # --- STEP 2: Load 551 ---
    file_551 = find_file("551")
    if file_551:
        df_551 = load_excel_as_str(file_551)
        if df_551 is not None:
            if all(col in df_551.columns for col in ['Pedimento', 'Fraccion', 'SecuenciaFraccion']):
                df_551['Pedimento_Key'] = (
                    df_551['Pedimento'] + 
                    df_551['Fraccion'] + 
                    df_551['SecuenciaFraccion']
                )
                
                map_551 = {
                    'ClaveDocumento': 'Clave Pedimento',
                    'ValorAduana': 'Valor Aduana',
                    'ValorComercial': 'Valor Comercial',
                    'DescripcionMercancia': 'Descripcion'
                }
                
                cols_to_pull = [c for c in map_551.keys() if c in df_551.columns]
                subset_551 = df_551[['Pedimento_Key'] + cols_to_pull].drop_duplicates('Pedimento_Key')
                df_557 = df_557.merge(subset_551, on='Pedimento_Key', how='left')
                df_557.rename(columns=map_551, inplace=True)

    # --- STEP 3: Load 505 ---
    file_505 = find_file("505")
    if file_505:
        df_505 = load_excel_as_str(file_505)
        if df_505 is not None and 'Pedimento' in df_505.columns:
            
            map_505 = {
                'ProveedorMercancia': 'Proveedor',
                'IndentFiscalProveedor': 'TaxID',
                'PaisFacturacion': 'País'
            }
            
            target_fields = list(map_505.keys())
            existing_fields = [c for c in target_fields if c in df_505.columns]
            
            # Clean TaxID (remove non-alphanumeric)
            if 'IndentFiscalProveedor' in df_505.columns:
                df_505['IndentFiscalProveedor'] = df_505['IndentFiscalProveedor'].apply(
                    lambda x: re.sub(r'[^a-zA-Z0-9]', '', str(x)) if pd.notna(x) else x
                )

            # Sort logic for deduplication
            df_505[existing_fields] = df_505[existing_fields].replace('', np.nan)
            df_505['data_score'] = df_505[existing_fields].notna().sum(axis=1)
            df_505_sorted = df_505.sort_values(by=['Pedimento', 'data_score'], ascending=[True, False])
            subset_505 = df_505_sorted[['Pedimento'] + existing_fields].drop_duplicates('Pedimento', keep='first')
            
            df_557['Pedimento_Clean'] = df_557['Pedimento']
            df_557 = df_557.merge(subset_505, left_on='Pedimento_Clean', right_on='Pedimento', how='left')
            df_557.rename(columns=map_505, inplace=True)
            
            if 'Pedimento_y' in df_557.columns: df_557.drop(columns=['Pedimento_y'], inplace=True)
            if 'Pedimento_x' in df_557.columns: df_557.rename(columns={'Pedimento_x': 'Pedimento'}, inplace=True)
            if 'Pedimento_Clean' in df_557.columns: df_557.drop(columns=['Pedimento_Clean'], inplace=True)

    # --- STEP 4: Load 701 (Rectified) & Apply Logic ---
    file_701 = find_file("701")
    rectified_in_col_b = set()
    rectified_in_col_f = set()
    
    if file_701:
        df_701 = load_excel_as_str(file_701)
        if df_701 is not None:
            if 'Pedimento' in df_701.columns:
                rectified_in_col_b.update(df_701['Pedimento'].dropna().unique())
            if 'PedimentoAnterior' in df_701.columns:
                rectified_in_col_f.update(df_701['PedimentoAnterior'].dropna().unique())

    # Logic:
    # rectificado = Pedimento ID if in 701 Col B, else 'NO'
    # anterior = Pedimento ID if in 701 Col F, else 'NO'
    
    def get_rectificado(ped):
        return ped if ped in rectified_in_col_b else 'NO'
    
    def get_anterior(ped):
        return ped if ped in rectified_in_col_f else 'NO'

    df_557['rectificado'] = df_557['Pedimento'].apply(get_rectificado)
    df_557['anterior'] = df_557['Pedimento'].apply(get_anterior)

    # --- Backfill Missing 'País' based on 'Proveedor' ---
    # Some rows might have Provider but missing Country. We check other rows with same Provider.
    if 'Proveedor' in df_557.columns and 'País' in df_557.columns:
        # Create a mapping of Provider -> Country (taking the first valid one found)
        # Filter rows where we have both
        valid_countries = df_557.dropna(subset=['Proveedor', 'País'])
        valid_countries = valid_countries[valid_countries['País'].str.strip() != '']
        
        if not valid_countries.empty:
            # Drop duplicates to get unique mappings. 
            # If a provider has multiple countries, keep the first one encountered.
            provider_country_map = valid_countries.set_index('Proveedor')['País'].to_dict()
            
            # Function to fill missing
            def fill_country(row):
                current_country = str(row['País']).strip()
                if not current_country or current_country.lower() == 'nan':
                    return provider_country_map.get(row['Proveedor'], row['País'])
                return row['País']

            df_557['País'] = df_557.apply(fill_country, axis=1)

    # --- Final Formatting ---
    target_cols = [
        'Patente', 'Pedimento', 'SeccionAduanera', 'Fraccion', 'SecuenciaFraccion', 
        'ClaveContribucion', 'FormaPago', 'ImportePago', 'FechaPagoReal', 
        'Clave Pedimento', 'Valor Aduana', 'Valor Comercial', 'Descripcion', 
        'Proveedor', 'TaxID', 'País', 'rectificado', 'anterior'
    ]
    
    for col in target_cols:
        if col not in df_557.columns:
            df_557[col] = ""
            
    df_final = df_557[target_cols].copy()
    
    # Fix Date Format: DD/MM/YYYY HH:MM -> YYYY-MM-DD HH:MM:SS
    if 'FechaPagoReal' in df_final.columns:
        df_final['FechaPagoReal'] = pd.to_datetime(df_final['FechaPagoReal'], dayfirst=True, errors='coerce')

    # Convert Numerics
    numeric_cols = ['Patente', 'Pedimento', 'SeccionAduanera', 'SecuenciaFraccion', 'ImportePago', 'Valor Aduana', 'Valor Comercial']
    for col in numeric_cols:
        if col in df_final.columns:
            df_final[col] = pd.to_numeric(df_final[col], errors='coerce')

    # --- Export ---
    output_file = "IVA_Importaciones.xlsx"
    print(f"Exporting to {output_file}...")
    
    try:
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df_final.to_excel(writer, index=False, sheet_name='IVA importaciones')
            ws = writer.sheets['IVA importaciones']
            
            header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            header_font = Font(bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
            
            rect_idx = df_final.columns.get_loc('rectificado') + 1
            for row in ws.iter_rows(min_row=2):
                # Highlight if rectificado != 'NO' (meaning it has an ID)
                if str(row[rect_idx-1].value) != 'NO':
                    for cell in row:
                        cell.fill = red_fill
                        
        print("Done.")
        
    except Exception as e:
        print(f"Error exporting: {e}")

if __name__ == "__main__":
    main()
