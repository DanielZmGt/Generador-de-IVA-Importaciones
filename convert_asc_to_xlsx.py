import pandas as pd
import glob
import os
from openpyxl.styles import Font, PatternFill

def detect_delimiter(filepath):
    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
            line = f.readline()
            if '|' in line: return '|'
            if '\t' in line: return '\t'
    except Exception:
        pass
    return '|'

def convert_asc_to_xlsx():
    asc_files = glob.glob("*.asc")
    if not asc_files:
        print("No .asc files found.")
        return

    print(f"Found {len(asc_files)} files to convert.")

    for filepath in asc_files:
        filename = os.path.basename(filepath)
        output_filename = os.path.splitext(filename)[0] + ".xlsx"
        print(f"Converting {filename} -> {output_filename}...")

        try:
            sep = detect_delimiter(filepath)
            # Use skipinitialspace=True to handle potential spaces after separators
            # index_col=False prevents pandas from treating the first col as index if there's a trailing separator misalignment
            df = pd.read_csv(filepath, sep=sep, dtype=str, encoding='utf-8', encoding_errors='ignore', skipinitialspace=True, index_col=False)
            
            # Clean up column names (strip whitespace)
            df.columns = df.columns.str.strip()
            
            # Clean up all data cells (strip whitespace)
            for col in df.columns:
                if df[col].dtype == 'object':
                    df[col] = df[col].str.strip()

            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Data')
                
                ws = writer.sheets['Data']
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
            
            print(f"Saved {output_filename}")

        except Exception as e:
            print(f"Error converting {filename}: {e}")

if __name__ == "__main__":
    convert_asc_to_xlsx()
